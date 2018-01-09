import discogs_client, os, time, statistics
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from urllib import request
os.chdir(r"D:\Dropbox\notes\dad's vinyls")

#Set up wd and load wb
#data_only parameter means don't read formulas
cwd = os.getcwd()
d = discogs_client.Client('ForMyDad/1.0', user_token="YcTxcautCkMMgvYAdSfAzfUZaleucoNGaniXupGP")
wb = load_workbook("dad's vinyls.xlsx", data_only = True)
ws = wb.get_sheet_by_name('Sheet1')

#Get number of catalog numbers
vinyl_count = ws['N2'].value

#Iterate through each row in spreadsheet
#D Album, E Artist, F Year, I sold median, J listings median, K checked
for i in range(2, vinyl_count + 2):
#for i in range(2, 6):
    try:
        if ws['K{}'.format(i)].value != 'Y':
            #To avoid getting shut out, sleep
            time.sleep(15)
            print('Looking at #{}'.format(i-1))
            #Read from wb
            catalog_number = ws['B{}'.format(i)].value
            
            #Read from api
            results = d.search(catno=catalog_number, type='master')
            results_release = d.search(catno=catalog_number, type='release')
            print('Length of master search results is {}'.format(len(results)))
            
            #Determine whether to use master or release in search
            if len(results) == 0:
                results = results_release
                is_master = 0
                print('Length of release search results is {}'.format(len(results)))               
            else:
                is_master = 1
                
            id_number = None
            id_number = results[0].id
            
            #Populate spreadsheet
            ws['C{}'.format(i)] = id_number
            ws['D{}'.format(i)] = results[0].title
            ws['L{}'.format(i)] = is_master
            try: #because masters often don't have artist or year
                ws['E{}'.format(i)] = results[0].artists[0].name
                ws['F{}'.format(i)] = results[0].year
            except Exception as e:
                print(e)                
                try:
                    ws['E{}'.format(i)] = results_release[0].artists[0].name
                    ws['F{}'.format(i)] = results_release[0].year
                except Exception as e:
                    print(e)
                    
            #Get median
            if is_master == 0:
                url = 'https://www.discogs.com/sell/release/{}?sort=price%2Casc&currency=USD&limit=250&ev=rb&page=1'.format(id_number)
            elif is_master == 1:            
                url = 'https://www.discogs.com/sell/list?sort=price%2Casc&limit=250&currency=USD&master_id={}&ev=mb'.format(id_number)
            
            print(id_number)
            print(url)     

            #Scrape selling page
            listings_median = None
            page = request.urlopen(url)
            ws['K{}'.format(i)] = 'Y'
            soup = BeautifulSoup(page, 'html.parser')
            listings_prices_all = soup.find_all('span', {'class':'price'})
            listings_prices_numeric = []
            for each in listings_prices_all:
                listings_prices_numeric.append(float(each.get_text()[1:]))

            listings_median = statistics.median(listings_prices_numeric)            
            ws['J{}'.format(i)] = listings_median
            

            
    
    except Exception as e:
        print(e)
    
    print('')
    
wb.save("dad's vinyls output.xlsx")

